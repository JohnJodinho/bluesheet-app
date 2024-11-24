from flask import (
    Flask, 
    jsonify, 
    request, 
    render_template, 
    send_file, 
    abort, 
    send_from_directory, 
    current_app,
    g
)

from werkzeug.utils import secure_filename
from PyPDF2 import PdfReader, PdfWriter

from prompt_templates import (
    SYSTEM_INSTRUCTION,
    prompt_template_misco, 
    prompt_template_shape_southwest, 

)

from typing import Iterable
import io
import time

import json
import os  # Handle file paths and directory operations
from os import path  # Check if files or directories exist, get file paths
import re  # Use regular expressions for text processing
from texts import SHAPE, SOUTH_WEST, MISCO
from queue import Queue
import threading
import requests
import logging

from openpyxl import Workbook
from openpyxl.styles import Font
from docx import Document as Docxdocument
from docx.shared import Pt


import google.generativeai as genai
from google.generativeai import (
    GenerativeModel,
    delete_file,
    configure,
    GenerationConfig, 
    ChatSession
)
from dotenv import load_dotenv


load_dotenv()
MODEL_NAME = os.getenv("MODEL_NAME")
BASE_URL = os.getenv("CLOUD_RUN_SERVICE_URL", "http://127.0.0.1:5000")
API_KEY = os.getenv("GENAI_API_KEY")



app = Flask(__name__)

configure(api_key=API_KEY)
# Set up logging
logging.basicConfig(
    filename="app.log",  # Log file in the root directory
    level=logging.DEBUG,  # Set the logging level (e.g., DEBUG, INFO, WARNING, ERROR, CRITICAL)
    format="%(asctime)s - %(levelname)s - %(message)s"  # Log message format
)

# Add a StreamHandler to also log to console
console_handler = logging.StreamHandler()
console_handler.setLevel(logging.DEBUG)
formatter = logging.Formatter("%(asctime)s - %(levelname)s - %(message)s")
console_handler.setFormatter(formatter)
app.logger.addHandler(console_handler)

# Configuration settings 
ALLOWED_EXTENSIONS = {'pdf'}

UPLOAD_FOLDER = 'uploads'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

TO_SPLIT_FOLDER = "to_split"
os.makedirs(TO_SPLIT_FOLDER, exist_ok=True)
app.config["TO_SPLIT_FOLDER"] = TO_SPLIT_FOLDER

TEMP_DOWNLOADS_FOLDER = 'downloads'
os.makedirs(TEMP_DOWNLOADS_FOLDER, exist_ok=True)
app.config['TEMP_DOWNLOADS_FOLDER'] = TEMP_DOWNLOADS_FOLDER

# Global variables
print_queue = Queue()
input_request_queue = Queue()
input_response_queue = Queue()
chat_thread = None
stop_chat = False
# Global variable to store error status and message
error_state = {"modelError": False, "message": ""}
batches_processed = 0


@app.route('/', methods=['GET'])
def home():
    global stop_chat, error_state
    stop_chat = False
    error_state = {"modelError": False, "message": ""}
    try:
        the_docs = os.listdir(app.config['UPLOAD_FOLDER'])
        uploaded = os.listdir(app.config["TO_SPLIT_FOLDER"])
        for the_doc in the_docs:
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], the_doc)
            if os.path.isfile(file_path):
                os.remove(file_path)
        uploaded_file = os.path.join(app.config["TO_SPLIT_FOLDER"], uploaded[0])
        if os.path.isfile(uploaded_file):
            os.remove(uploaded_file)
    except Exception as e:
        app.logger.error(f"Error removing file: {e}")
    app.logger.info("Home route accessed")
    return render_template('home.html')


@app.route('/chat', methods=['GET'])
def chat():
    # Reinitialize chat each time this route is accessed
    initialize_chat(reset=True)
    try:
        the_docs = os.listdir(app.config['UPLOAD_FOLDER'])
        uploaded = os.listdir(app.config["TO_SPLIT_FOLDER"])
        for the_doc in the_docs:
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], the_doc)
            if os.path.isfile(file_path):
                os.remove(file_path)
        uploaded_file = os.path.join(app.config["TO_SPLIT_FOLDER"], uploaded[0])
        if os.path.isfile(uploaded_file):
            os.remove(uploaded_file)
    except Exception as e:
        app.logger.error(f"Error removing file: {e}")
    app.logger.info("Chat route accessed")
    return render_template('chat.html')

def initialize_chat(reset=False):
    """Initialize or reset chat components."""
    global chat_thread, stop_chat, error_state
    if reset or not hasattr(g, 'chat_initialized'):
        # Signal any existing chat thread to stop
        stop_chat = True
        if chat_thread and chat_thread.is_alive():
            chat_thread.join()  # Wait for the thread to end
            # time.sleep(2)
        # Reset the control flag for the new session
        stop_chat = False
        error_state = {"modelError": False, "message": ""}
        chat_thread = threading.Thread(target=start_chat_session, daemon=True)
        chat_thread.start()
        g.chat_initialized = True

def check_and_delete_all_files():
    try:
        all_files = genai.list_files()
        for f in all_files:
            f.delete()
    except Exception as e:
        app.logger.warning(f"Could not delete files: {e}")

def start_chat_session():
    """Main chat session logic"""
    
    time.sleep(1)  
    
    
    model = GenerativeModel(
        MODEL_NAME,
        system_instruction=SYSTEM_INSTRUCTION
    )

    session = model.start_chat()
    check_and_delete_all_files()
    handle_step_one(session)
    handle_step_two(session)
    handle_last_step(session)
    check_and_delete_all_files()
    
    try:
        the_docs = os.listdir(app.config['UPLOAD_FOLDER'])
        uploaded = os.listdir(app.config["TO_SPLIT_FOLDER"])
        for the_doc in the_docs:
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], the_doc)
            if os.path.isfile(file_path):
                os.remove(file_path)
        uploaded_file = os.path.join(app.config["TO_SPLIT_FOLDER"], uploaded[0])
        if os.path.isfile(uploaded_file):
            os.remove(uploaded_file)
    except Exception as e:
        app.logger.error(f"Error removing file: {e}")


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS  


@app.route('/upload', methods=['POST'])
def upload_file():
    global batches_processed
    if 'file' not in request.files:
        return jsonify({"error": "No file part in the request"}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({"error": "No file selected for uploading"}), 400

    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)  # Secure the filename
        save_path = os.path.join(app.config['TO_SPLIT_FOLDER'], filename)

        # Save the file to the upload folder
        file.save(save_path)

        # Call the utility function with the saved file path
        try:
            batches_processed = split_and_save_pdf(file_name=save_path)
            response_data = {
                "filename": filename,
                "content_preview": "PDF split successfully"
            }
            return jsonify(response_data), 200
        except Exception as e:
            return jsonify({"error": str(e)}), 500

    return jsonify({"error": "Invalid file type"}), 400

# Endpoint to list new files in the downloads folder
@app.route('/check-new-files', methods=['GET'])
def check_new_files():
    try:
        files = os.listdir(app.config['TEMP_DOWNLOADS_FOLDER'])
        return jsonify(files=files), 200
    except Exception as e:
        app.logger.error(f"Error checking new files: {e}")
        return jsonify({"error": str(e)}), 500

@app.route("/download/<path:filename>", methods=["GET"])
def download(filename):
    try:
        downloads = os.path.join(current_app.root_path, app.config['TEMP_DOWNLOADS_FOLDER'])
        
        file_path = os.path.join(downloads, filename)
        app.logger.info(f"Attempting to download: {file_path}")

        # Check if the file exists
        if os.path.isfile(file_path):
            return send_from_directory(downloads, filename, as_attachment=True)
        else:
            abort(404, description="File not found")
    except Exception as e:
        app.logger.error(f"Error downloading file: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/remove/<filename>', methods=['DELETE'])
def remove_file(filename):
    try:
        file_path = os.path.join(app.config['TEMP_DOWNLOADS_FOLDER'], filename)
        if os.path.isfile(file_path):
            os.remove(file_path)
            return jsonify({"message": "File deleted successfully"}), 200
        else:
            return jsonify({"error": "File not found"}), 404
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/message', methods=['POST'])
def message_handler():
    action = request.json.get('action')

    if action == "print":
        # Queue a message for the client
        message = request.json.get('message')
        print_queue.put(message)
        return jsonify({"status": "Message queued for client"}), 200

    elif action == "input_request":
        # Queue an input request
        input_request_queue.put("Input requested")
        return jsonify({"status": "Input request sent to client"}), 200

    elif action == "input_response":
        # Store client input in response queue
        response = request.json.get('response')
        if response is not None:
            input_response_queue.put(response)
            return jsonify({"status": "Input received from client"}), 200
        return jsonify({"error": "No response provided"}), 400

    elif action == "fetch":
        
        messages = []
        while not print_queue.empty():
            messages.append(print_queue.get())

        input_prompt = None
        if not input_request_queue.empty():
            input_prompt = input_request_queue.get()

        return jsonify({"messages": messages, "input_prompt": input_prompt}), 200

    return jsonify({"error": "Invalid action"}), 400


@app.route('/get_error_state', methods=['GET'])
def get_error_state():
    # Send the current error state to the front end
    return jsonify(error_state)

def custom_print(message):
    requests.post(f"{BASE_URL}/message", json={"action": "print", "message": message})

# Function to request client input, waits for a response
def custom_input():
    requests.post(f"{BASE_URL}/message", json={"action": "input_request"})

    # Wait for the client to respond
    while True:
        if not input_response_queue.empty():
            return input_response_queue.get()
        time.sleep(1)  # Poll every second
def check_and_delete_all_files():
    try:
        all_files = genai.list_files()
        for f in all_files:
            f.delete()
    except Exception as e:
        app.logger.warning(f"Could not delete files: {e}")


def generate(
    prompt: list,
    max_output_tokens: int = 2048,
    temperature: int = 2,
    top_p: float = 0.4,
    stream: bool = False,
    session = None
):
    """
    Function to generate response using Gemini 1.5 Pro
    """
    global stop_chat  # Refer to the global stop_chat flag
    global error_state  # Refer to the global error state
    stage = 0  # Initialize the stage counter
    while not stop_chat:
        try:
            
            responses = session.send_message(
                prompt,
                generation_config={
                    "max_output_tokens": max_output_tokens,
                    "temperature": temperature,
                    "top_p": top_p,
                },
                
                stream=stream,
            )
            error_state = {"modelError": False, "message": ""}
            
            break  # Exit loop if no error occurs
        except Exception as e:
            app.logger.error(f"An error occurred, retrying in 2 seconds: {e}")
            if "429" in str(e):
                continue
            message = f"Gemini Model Error: {str(e)}"
            if stage == 0:
                error_state = {"modelError": True, "message": message}
                stage += 1
            time.sleep(3)

    return responses if not stop_chat else None  # Return None if stopped

def is_valid_json(response_text):
    """
    Check if response_text is valid JSON.
    
    Strips any JSON code block delimiters (e.g., ```json) and 
    then attempts to parse it to confirm validity.
    """
    # Remove JSON code block delimiters if they exist
    json_text = re.sub(r'```json|```', '', response_text).strip()

    try:
        # Try loading the text as JSON
        json.loads(json_text)
        return True
    except ValueError:
        # If parsing fails, it's not valid JSON
        return False

def clean_html(response_text):
    html = re.sub(r'```html|```', '', response_text).strip()
    return html

def save_to_excel(json_data, folder=None):
    """Save JSON data to Excel with bold headers using openpyxl."""
    json_text = re.sub(r'```json|```', '', json_data).strip()
    json_data = json.loads(json_text)
    wb = Workbook()
    document_name = json_data.get("document_name", "Bluesheet Draft.xlsx")
    
    default_sheet = wb.active
    wb.remove(default_sheet)
    for sheet_name, rows in json_data.items():
        if sheet_name == "document_name":
            continue  
        
        ws = wb.create_sheet(title=sheet_name)

        if rows:

            headers = rows[0].keys()
            ws.append(list(headers))
            
            for cell in ws[1]:  # First row (headers)
                cell.font = Font(bold=True)
            
            for row_data in rows:
                ws.append(list(row_data.values()))
        else:
            app.logger.warning(f"Warning: No data found for sheet '{sheet_name}'.")

    # Ensure the TEMP_DOWNLOADS_FOLDER is set in the config
    folder = app.config.get('TEMP_DOWNLOADS_FOLDER')
    if not folder:
        app.logger.critical("TEMP_DOWNLOADS_FOLDER is not configured in app.config.")
    
    # Ensure the folder exists
    os.makedirs(folder, exist_ok=True)
    save_path = os.path.join(folder, document_name)
    
    # Save the workbook
    wb.save(save_path)
    app.logger.info(f"Document saved as {save_path}")
    

    

def split_and_save_pdf(file_name):
    input_pdf = PdfReader(file_name)
    batch_size = 50
    num_batches = len(input_pdf.pages) // batch_size + 1

    # Extract batches of pages from the PDF
    for b in range(num_batches):
        writer = PdfWriter()

        # Get the start and end page numbers for this batch
        start_page = b * batch_size
        end_page = min((b + 1) * batch_size, len(input_pdf.pages))

        # Add pages in this batch to the writer
        for i in range(start_page, end_page):
            writer.add_page(input_pdf.pages[i])

        # Save the batch to a separate PDF file
        batch_filename = f'{os.path.splitext(os.path.basename(file_name))[0]}-batch{b+1}.pdf'
        file_path = os.path.join(UPLOAD_FOLDER, batch_filename)
        with open(file_path, 'wb') as output_file:
            writer.write(output_file)
    return num_batches



def load_document(file_path, folder=None):
    file_extension = file_path.split('.')[-1].lower()
    _mime_type = None
    if file_extension == "pdf":
        _mime_type = "application/pdf"
    elif file_extension == "csv":
        file_path = csv_to_txt(file_path, folder)
        _mime_type = "text/plain"
    elif file_extension == "txt":
        _mime_type = "text/plain"
    elif file_extension in ["doc", "docx"]:
        word_to_pdf(file_path, folder)
        file_path = f"{folder}/{file_path.replace('.docx', '.pdf')}"
        _mime_type = "application/pdf"
    else:
        app.logger.warning("Unsupported file type. Please provide a PDF, or TXT document.") 
    # Load the file
    with open(file_path, "rb") as f:
        document = genai.upload_file(f, mime_type=_mime_type)
    return document

def handle_step_one(session):
    stage = 0
    while not stop_chat:
        rfp_docs = None
        if stage == 0: 
            custom_print("Please upload the RFP document to generate the Bluesheet.")
            stage +=1
        while rfp_docs == None and not stop_chat:
            try:
                rfp_docs = os.listdir(app.config['UPLOAD_FOLDER'])
                if batches_processed != 0 and len(rfp_docs) == batches_processed:   
                    # print(len(rfp_docs))
                    rfp_docs = [os.path.join(UPLOAD_FOLDER, rfp_doc) for rfp_doc in rfp_docs]
                else:
                    rfp_docs = None
               
                if len(rfp_docs) < 1:
                    rfp_docs = None
            except Exception as e:
                # app.logger.error(f"Error in handle_step_one: No file exists in {UPLOAD_FOLDER}. Error: {e}")
                rfp_docs = None
                

        rfp_part_files = {}
        
        custom_print("Getting the RFP document ready...")
        for rfp_document_name in rfp_docs:
            rfp_document = load_document(rfp_document_name)
            rfp_part_files[rfp_document_name] = rfp_document
    
        rfp_parts = sorted_dict = dict(
            sorted(rfp_part_files.items(), key=lambda item: int(item[0].split("-batch")[1].split(".pdf")[0]))
        )

        rfp_parts_prompt = list(rfp_parts.values())
        rfp_parts_prompt.pop(len(rfp_parts_prompt)-1)
      
        rfp_parts_prompt.append(f"Here is the rfp document splitted into {len(rfp_parts_prompt)} chunks (.pdf files) for easy processing")
        custom_print("Analyzing the RFP document...")
        # Request JSON content for the document

        response = generate(prompt=rfp_parts_prompt, session=session)
        if response == None:
            return
        # custom_print(response.text)
        custom_print("Generating Bluesheet for MISCO...")
        try:
            response_text = generate(prompt=[prompt_template_misco, MISCO], session=session).text
        except Exception as e:
            app.logger.error(f"Error ocurred in handle_step_one: {e}")
            return
        # custom_print(response_text)
        if is_valid_json(response_text):
            try:
                # Save or pass the JSON data to the json_to_docx function
                print(response_text)
                save_to_excel(response_text, TEMP_DOWNLOADS_FOLDER)
                custom_print("Bluesheet document successfuly generated!")
                break
            except Exception as e: 
                app.logger.error(f"An error occurred in handle_step_one: {e}")
        
        app.logger.error("Error occurred while generating the JSON structure. Trying again...")
        custom_print("An error occurred")
        custom_print("Regenerating Bluesheet Document...")
        try:
            responses = generate(prompt=[
            "The generation failed. Please try again, ensuring no errors in JSON structure."
            ], session=session)
            response_text = responses.text
        except Exception as e:
            app.logger.error(f"Error ocurred in handle_step_one: {e}")
        
        if is_valid_json(response_text):
            try:
                save_to_excel(response_text, TEMP_DOWNLOADS_FOLDER)
                custom_print("Bluesheet document successfuly generated!")
                break

            except Exception as e: 
                app.logger.error(f"An error occurred handle_step_one: {e}")
                custom_print("An error ocurred. Try again later")
                return 
    # Handle user feedback and modifications
    while not stop_chat:
        custom_print("Would you like to make any modifications or corrections to the document? If yes, please specify the modifications.")
        user_response = custom_input()
        # Request JSON content for the document
        try:
            response_text = generate(prompt=[
                f"""
                <NOTE> This stage is iterative depending on the user's feedback in <USER_RESPONSE> </NOTE>
                <INSTRUCTIONS>
                Apply user feedback to the previously sent JSON and send new with changes. That is if user requests modifications.
                <ACTION>
                If the user requests modifications:
                    * apply the user feedback to the JSON document and return the JSON that'll be formatted into a .xlsx document.
                If the user does not request modifications:
                    * Reply with a message indicating that no modifications were requested and proceeding to the next step. (Message should not be in JSON format)
                    * Do not mention anything about JSON in your response.
                <USER_RESPONSE>{user_response}</USER_RESPONSE>
                """
            ],  session=session).text
            if is_valid_json(response_text):
                try:
                    save_to_excel(response_text, TEMP_DOWNLOADS_FOLDER)
                    custom_print("Document updated successfully.")
                    continue
                except Exception as e:
                    app.logger.error(f"An error occurred: {e}")
                    continue
            else: 
                custom_print(clean_html(response_text))
                return
        except Exception as e:
            app.logger.info(f"An error occured in handle_step_one: {e}")
            return

def handle_step_two(session):
    custom_print("""
Would you like to identify synergy opportunities in this RFP for companies:  
Shape (representing related companies in Northern California)  
and Southwest Valve? I will review the RFP based on their products and categories and generate the bluesheet. 
""")
    while not stop_chat:
        user_response = custom_input()
        prompt_template_shape_southwest.format(user_response=user_response)
        custom_print("Generating Bluesheet for Shape and Southwest...")
        try:
            response_text = generate(prompt=[prompt_template_shape_southwest, SHAPE, SOUTH_WEST], session=session).text
        except Exception as e:
            app.logger.error(f"Error occurred in handle_step_two: {e}")
            return
        if is_valid_json(response_text):
            try:
                # Save or pass the JSON data to the json_to_docx function
                print(response_text)
                save_to_excel(response_text, TEMP_DOWNLOADS_FOLDER)
                custom_print("Bluesheet document successfuly generated.")
                break
            except Exception as e: 
                app.logger.error(f"An error occurred: {e}")
        else:
            custom_print(clean_html(response_text))
            return
        
        app.logger.error("Error occurred while generating the JSON structure. Trying again...")
        custom_print("An error occurred")
        custom_print("Regenerating Bluesheet Document...")
        responses = generate(prompt=[
            """The document generation failed. Please try again, ensuring no errors in JSON structure."""
        ], session=session)
        if responses == None:
            return
        response_text = responses.text
        if is_valid_json(response_text):
            try:
                
                save_to_excel(response_text, TEMP_DOWNLOADS_FOLDER)
                custom_print("Document successfuly generated")
                break
                
            except Exception as e: 
                app.logger.error(f"An error occurred: {e}")
                custom_print("An error ocurred. Try again later")
                return False
    # Handle user feedback and modifications
    while not stop_chat:
        custom_print("Would you like to make any modifications or corrections to the document? If yes, please specify the modifications.")
        user_response = custom_input().strip()
        # Request JSON content for the document
        try:
            response_text = generate(prompt=[
                f"""
                <NOTE> This stage is iterative depending on the user's feedback in <USER_RESPONSE> </NOTE>
                <INSTRUCTIONS>
                Apply user feedback to the previously sent JSON and send new with changes. That is if user requests modifications.
                <ACTION>
                If the user requests modifications:
                    * apply the user feedback to the JSON document and return the JSON that'll be formatted into a .xlsx document.
                If the user does not request modifications:
                    * Reply with a message indicating that no modifications were requested and proceeding to the next step. (Message should not be in JSON format)
                    * Do not mention anything about JSON in your response.
                <USER_RESPONSE>{user_response}</USER_RESPONSE>
                """
            ],  session=session).text
            if is_valid_json(response_text):
                try:
                    save_to_excel(response_text, TEMP_DOWNLOADS_FOLDER)
                    custom_print("Document updated successfully.")
                    continue
                except Exception as e:
                    app.logger.error(f"An error occurred: {e}")
                    continue
            else: 
                custom_print(clean_html(response_text))
                return
        except Exception as e:
            return

def handle_last_step(session):
    while not stop_chat:
        custom_print("Would you like to perform any other actions? (yes/no)")
        user_response = custom_input()
        if user_response.lower().find("yes") != -1:
            custom_print("Provide details on the next action you would like to take, and I will assist you. or type 'end' to end this session")
            user_response = custom_input()
            if user_response.lower().find("end") != -1: 
                custom_print("I will conclude the session. Thank you for using this service.")
                break
            else:
                response = generate(prompt=[user_response], session=session).text
        else:
            custom_print("I will conclude the session. Thank you for using this service.")
            break


        
        

def run_server():
    """Server initialization function"""
    if BASE_URL == "http://127.0.0.1:5000":
        app.run(debug=True, use_reloader=False)
    else:
        port = 8080
        app.run(host="0.0.0.0", port=port, debug=False, use_reloader=False)

if __name__ == "__main__":
    # Start only the Flask server initially
    
    run_server()

